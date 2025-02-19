import tkinter as tk
from tkinter import messagebox, ttk, filedialog, simpledialog
import sqlite3
import re
from openpyxl import Workbook
from openpyxl import load_workbook

# Создание базы данных и таблиц
def create_database():
    conn = sqlite3.connect('auto_parts.db')
    c = conn.cursor()

    # Таблица для автозапчастей
    c.execute('''
        CREATE TABLE IF NOT EXISTS parts (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            price REAL NOT NULL,
            quantity INTEGER NOT NULL
        )
    ''')

    # Таблица для клиентов
    c.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            phone TEXT NOT NULL
        )
    ''')

    # Таблица для продаж
    c.execute('''
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY,
            client_id INTEGER NOT NULL,
            part_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            sale_date TEXT NOT NULL,
            FOREIGN KEY (client_id) REFERENCES clients(id),
            FOREIGN KEY (part_id) REFERENCES parts(id)
        )
    ''')

    # Таблица для автомобилей
    c.execute('''
        CREATE TABLE IF NOT EXISTS cars (
            id INTEGER PRIMARY KEY,
            make TEXT NOT NULL,
            model TEXT NOT NULL,
            year INTEGER NOT NULL
        )
    ''')

    # Таблица для поставщиков
    c.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            contact TEXT NOT NULL
        )
    ''')

    # Таблица для заказов
    c.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY,
            supplier_id INTEGER NOT NULL,
            part_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            order_date TEXT NOT NULL,
            status TEXT DEFAULT 'В обработке',
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY (part_id) REFERENCES parts(id)
        )
    ''')

    # Проверка наличия колонки `status` в таблице `orders`
    c.execute("PRAGMA table_info(orders)")
    columns = [column[1] for column in c.fetchall()]
    if "status" not in columns:
        c.execute("ALTER TABLE orders ADD COLUMN status TEXT DEFAULT 'В обработке'")

    # Проверка наличия колонки `category` в таблице `parts`
    c.execute("PRAGMA table_info(parts)")
    columns = [column[1] for column in c.fetchall()]
    if "category" not in columns:
        c.execute("ALTER TABLE parts ADD COLUMN category TEXT DEFAULT 'Общая'")

    conn.commit()
    conn.close()

# Главное окно
class AutoPartsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("АРМ Продавца Автозапчастей")
        self.root.geometry("1000x600")
        self.root.configure(bg="#f9f9f9")

        # Маппинг между русскими названиями и именами таблиц
        self.table_name_mapping = {
            "автозапчастей": "parts",
            "клиентов": "clients",
            "автомобилей": "cars",
            "поставщиков": "suppliers",
            "заказов": "orders"
        }

        # Маппинг между английскими названиями колонок и русскими заголовками
        self.column_headers = {
            "parts": {
                "ID": "ID",
                "name": "Наименование",
                "category": "Категория",
                "price": "Цена",
                "quantity": "Количество"
            },
            "clients": {
                "ID": "ID",
                "name": "ФИО",
                "phone": "Телефон"
            },
            "cars": {
                "ID": "ID",
                "make": "Марка",
                "model": "Модель",
                "year": "Год выпуска"
            },
            "suppliers": {
                "ID": "ID",
                "name": "Название",
                "contact": "Контакты"
            },
            "orders": {
                "ID": "ID",
                "Supplier": "Поставщик",
                "Part": "Запчасть",
                "Quantity": "Количество",
                "Date": "Дата",
                "Status": "Статус"
            }
        }

        # Закрепленная верхняя панель
        self.create_menu()

        # Основной контейнер для отображения содержимого
        self.main_frame = tk.Frame(self.root, bg="#ffffff", padx=20, pady=20)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Начальный фрейм
        self.start_frame = tk.Frame(self.main_frame, bg="#ffffff", padx=20, pady=20)
        self.start_frame.pack(fill=tk.BOTH, expand=True)

        label = tk.Label(self.start_frame, text="Добро пожаловать в АРМ Продавца Автозапчастей!", font=("Arial", 20, "bold"), bg="#ffffff")
        label.pack(pady=50)

    # Создание закрепленного меню
    def create_menu(self):
        menu_bar = tk.Menu(self.root, bg="#ffffff", fg="#333333", activebackground="#4CAF50",
                           activeforeground="#ffffff")
        self.root.config(menu=menu_bar)

        # Справочники
        ref_menu = tk.Menu(menu_bar, tearoff=0, bg="#ffffff", fg="#333333", activebackground="#4CAF50",
                           activeforeground="#ffffff")
        ref_menu.add_command(label="Автозапчасти", command=lambda: self.show_ref("parts"))
        ref_menu.add_command(label="Клиенты", command=lambda: self.show_ref("clients"))
        ref_menu.add_command(label="Автомобили", command=lambda: self.show_ref("cars"))
        ref_menu.add_command(label="Поставщики", command=lambda: self.show_ref("suppliers"))
        ref_menu.add_command(label="Заказы", command=lambda: self.show_ref("orders"))
        ref_menu.add_command(label="Импорт из Excel", command=self.import_from_excel)  # Новый пункт
        menu_bar.add_cascade(label="Справочники", menu=ref_menu)

        # Отчеты
        report_menu = tk.Menu(menu_bar, tearoff=0, bg="#ffffff", fg="#333333", activebackground="#4CAF50",
                              activeforeground="#ffffff")
        report_menu.add_command(label="Продажи", command=self.show_sales_report)
        report_menu.add_command(label="Заказы", command=self.show_orders_report)
        report_menu.add_command(label="Экспорт в Excel", command=self.export_to_excel)
        menu_bar.add_cascade(label="Отчеты", menu=report_menu)

        # Формирование заказа
        menu_bar.add_command(label="Создать заказ", command=self.create_order, background="#4CAF50",
                             foreground="#ffffff", activebackground="#388E3C", activeforeground="#ffffff")

        # Руководство
        menu_bar.add_command(label="Руководство", command=self.show_manual, background="#FF9800", foreground="#ffffff",
                             activebackground="#F57C00", activeforeground="#ffffff")

    # Показать справочник
    def show_ref(self, table_name):
        self.clear_main_frame()

        frame = tk.Frame(self.main_frame, bg="#ffffff", padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        title_map = {
            "parts": "Справочник автозапчастей",
            "clients": "Справочник клиентов",
            "cars": "Справочник автомобилей",
            "suppliers": "Справочник поставщиков",
            "orders": "Список заказов"
        }

        label = tk.Label(frame, text=title_map.get(table_name, ""), font=("Arial", 18, "bold"), bg="#ffffff")
        label.pack(pady=10)

        tree = self.create_treeview(frame, table_name)
        self.load_data(tree, table_name)

        if table_name == "orders":
            # Кнопка для обновления статуса
            button_frame = tk.Frame(frame, bg="#ffffff")
            button_frame.pack(pady=10)

            update_status_button = tk.Button(button_frame, text="Обновить статус", font=("Arial", 12), bg="#FFC107", fg="#ffffff", command=lambda: self.update_order_status(tree))
            update_status_button.pack(side=tk.LEFT, padx=10)

            # Кнопка для редактирования заказа
            edit_order_button = tk.Button(button_frame, text="Редактировать", font=("Arial", 12), bg="#2196F3", fg="#ffffff", command=lambda: self.edit_order(tree))
            edit_order_button.pack(side=tk.LEFT, padx=10)

        elif table_name != "orders":
            # Кнопки для добавления, редактирования и удаления
            button_frame = tk.Frame(frame, bg="#ffffff")
            button_frame.pack(pady=10)

            add_button = tk.Button(button_frame, text="Добавить", font=("Arial", 12), bg="#4CAF50", fg="#ffffff", command=lambda: self.add_record(table_name))
            add_button.pack(side=tk.LEFT, padx=10)

            edit_button = tk.Button(button_frame, text="Редактировать", font=("Arial", 12), bg="#2196F3", fg="#ffffff", command=lambda: self.edit_record(tree, table_name))
            edit_button.pack(side=tk.LEFT, padx=10)

            delete_button = tk.Button(button_frame, text="Удалить", font=("Arial", 12), bg="#FF5722", fg="#ffffff", command=lambda: self.delete_record(tree, table_name))
            delete_button.pack(side=tk.LEFT, padx=10)

    # Создание Treeview для справочников
    def create_treeview(self, frame, table_name):
        columns_map = {
            "parts": ("ID", "name", "category", "price", "quantity"),
            "clients": ("ID", "name", "phone"),
            "cars": ("ID", "make", "model", "year"),
            "suppliers": ("ID", "name", "contact"),
            "orders": ("ID", "Supplier", "Part", "Quantity", "Date", "Status")
        }

        # Получаем список колонок для текущей таблицы
        columns = columns_map.get(table_name, ())
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=15, style="Custom.Treeview")

        # Устанавливаем русские заголовки
        for col in columns:
            russian_header = self.column_headers[table_name].get(col, col)
            tree.heading(col, text=russian_header)
            tree.column(col, width=150)

        tree.pack(fill=tk.BOTH, expand=True)

        # Чередование цветов строк
        tree.tag_configure("oddrow", background="#f0f0f0")
        tree.tag_configure("evenrow", background="#ffffff")

        return tree

    # Загрузка данных в Treeview
    def load_data(self, tree, table_name):
        tree.delete(*tree.get_children())
        conn = sqlite3.connect('auto_parts.db')
        c = conn.cursor()

        if table_name == "orders":
            c.execute("""
                SELECT orders.id, suppliers.name, parts.name, orders.quantity, orders.order_date, orders.status
                FROM orders
                INNER JOIN suppliers ON orders.supplier_id = suppliers.id
                INNER JOIN parts ON orders.part_id = parts.id
            """)
        else:
            c.execute(f"SELECT * FROM {table_name}")

        rows = c.fetchall()
        for i, row in enumerate(rows):
            tag = "oddrow" if i % 2 == 0 else "evenrow"
            tree.insert("", tk.END, values=row, tags=(tag,))
        conn.close()

    # Добавление новой записи
    def add_record(self, table_name):
        add_window = tk.Toplevel(self.root)
        add_window.title(f"Добавление записи ({table_name})")
        add_window.geometry("400x300")
        add_window.configure(bg="#f9f9f9")

        fields_map = {
            "parts": ["name", "category", "price", "quantity"],
            "clients": ["name", "phone"],
            "cars": ["make", "model", "year"],
            "suppliers": ["name", "contact"]
        }

        fields = fields_map.get(table_name, [])
        entries = []

        input_frame = tk.Frame(add_window, bg="#f9f9f9")
        input_frame.pack(pady=10)

        for i, field in enumerate(fields):
            label = tk.Label(input_frame, text=self.column_headers[table_name].get(field, field) + ":", font=("Arial", 12), bg="#f9f9f9")
            label.grid(row=i, column=0, sticky="w", padx=10, pady=5)

            if field == "category":
                category_var = tk.StringVar()
                category_combo = ttk.Combobox(input_frame, textvariable=category_var, values=["Двигатель", "Тормозная система", "Электрика"], font=("Arial", 12))
                category_combo.grid(row=i, column=1, padx=10, pady=5)
                entries.append(category_var)
            else:
                entry = tk.Entry(input_frame, font=("Arial", 12))
                entry.grid(row=i, column=1, padx=10, pady=5)
                entries.append(entry)

        def save_record():
            new_values = [entry.get() if isinstance(entry, tk.Entry) else entry.get() for entry in entries]
            print("Введенные данные:", new_values)  # Отладочное сообщение

            if any(not value for value in new_values):
                messagebox.showerror("Ошибка", "Все поля должны быть заполнены!")
                return

            if not self.validate_data(new_values, table_name):
                return

            conn = sqlite3.connect('auto_parts.db')
            c = conn.cursor()
            insert_query = f"INSERT INTO {table_name} (" + ", ".join(fields) + ") VALUES (" + ", ".join(["?"] * len(fields)) + ")"
            c.execute(insert_query, new_values)
            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Запись успешно добавлена!")
            add_window.destroy()
            self.refresh_current_view()

        button_frame = tk.Frame(add_window, bg="#f9f9f9")
        button_frame.pack(pady=10)

        save_button = tk.Button(button_frame, text="Сохранить", font=("Arial", 12), bg="#4CAF50", fg="#ffffff", command=save_record)
        save_button.pack(side=tk.LEFT, padx=10)

        cancel_button = tk.Button(button_frame, text="Отмена", font=("Arial", 12), bg="#FF5722", fg="#ffffff", command=add_window.destroy)
        cancel_button.pack(side=tk.LEFT, padx=10)

    # Валидация данных
    def validate_data(self, data, table_name):
        if table_name == "parts":
            try:
                float(data[2])  # Цена
                int(data[3])    # Количество
            except ValueError:
                messagebox.showerror("Ошибка", "Цена должна быть числом, количество — целым числом!")
                return False
        elif table_name == "clients":
            if not re.match(r"^\+?\d{10,15}$", data[1]):
                messagebox.showerror("Ошибка", "Неверный формат телефона!")
                return False
        elif table_name == "cars":
            try:
                int(data[2])  # Год выпуска
            except ValueError:
                messagebox.showerror("Ошибка", "Год должен быть целым числом!")
                return False
        return True

    # Экспорт данных в Excel
    def export_to_excel(self):
        # Выбор таблицы для экспорта
        table_name = "parts"  # Можно сделать выбор через диалоговое окно

        # Подключение к базе данных
        conn = sqlite3.connect('auto_parts.db')
        c = conn.cursor()

        # Загрузка данных из таблицы
        c.execute(f"SELECT * FROM {table_name}")
        rows = c.fetchall()

        # Получение заголовков колонок
        headers = list(self.column_headers[table_name].values())

        # Создание новой книги Excel
        wb = Workbook()
        ws = wb.active
        ws.title = table_name.capitalize()  # Название листа

        # Запись заголовков
        ws.append(headers)

        # Запись данных
        for row in rows:
            ws.append(row)

        # Сохранение файла
        file_path = f"{table_name}.xlsx"
        wb.save(file_path)
        conn.close()

        messagebox.showinfo("Успех", f"Данные успешно экспортированы в файл: {file_path}")

    # Экспорт данных в Excel с выбором таблицы
    def export_to_excel(self):
        # Диалоговое окно для выбора таблицы
        table_name = tk.simpledialog.askstring("Выбор таблицы", "Введите название таблицы (например, parts, clients):")
        if not table_name or table_name not in self.table_name_mapping.values():
            messagebox.showerror("Ошибка", "Неверное название таблицы!")
            return

        # Подключение к базе данных
        conn = sqlite3.connect('auto_parts.db')
        c = conn.cursor()

        # Загрузка данных из таблицы
        c.execute(f"SELECT * FROM {table_name}")
        rows = c.fetchall()

        # Получение заголовков колонок
        headers = list(self.column_headers[table_name].values())

        # Создание новой книги Excel
        wb = Workbook()
        ws = wb.active
        ws.title = table_name.capitalize()  # Название листа

        # Запись заголовков
        ws.append(headers)

        # Запись данных
        for row in rows:
            ws.append(row)

        # Сохранение файла
        file_path = f"{table_name}.xlsx"
        wb.save(file_path)
        conn.close()

        messagebox.showinfo("Успех", f"Данные успешно экспортированы в файл: {file_path}")

    # Импорт данных из Excel
    def import_from_excel(self):
        # Запрос пути к файлу Excel
        file_path = filedialog.askopenfilename(
            title="Выберите файл Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            messagebox.showwarning("Внимание", "Файл не выбран!")
            return

        # Запрос названия таблицы
        table_name = tk.simpledialog.askstring("Выбор таблицы", "Введите название таблицы (например, parts, clients):")
        if not table_name or table_name not in self.table_name_mapping.values():
            messagebox.showerror("Ошибка", "Неверное название таблицы!")
            return

        # Чтение данных из Excel
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # Получение заголовков из первой строки
            headers = [cell.value for cell in ws[1]]

            # Проверка соответствия заголовков
            expected_headers = list(self.column_headers[table_name].values())
            if headers != expected_headers:
                messagebox.showerror("Ошибка", "Заголовки в файле не соответствуют ожидаемым!")
                return

            # Подключение к базе данных
            conn = sqlite3.connect('auto_parts.db')
            c = conn.cursor()

            # Очистка таблицы перед импортом (опционально)
            confirm = messagebox.askyesno("Подтверждение", "Очистить таблицу перед импортом?")
            if confirm:
                c.execute(f"DELETE FROM {table_name}")

            # Чтение данных из Excel и запись в базу данных
            for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем первую строку (заголовки)
                insert_query = f"INSERT INTO {table_name} (" + ", ".join(headers) + ") VALUES (" + ", ".join(
                    ["?"] * len(headers)) + ")"
                c.execute(insert_query, row)

            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Данные успешно импортированы!")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при импорте: {str(e)}")

    # Обновление текущего представления
    def refresh_current_view(self):
        current_frame = self.main_frame.winfo_children()[0]
        children = current_frame.winfo_children()

        title_label = None
        for widget in children:
            if isinstance(widget, tk.Label):
                title_label = widget
                break

        if not title_label or "Справочник" not in title_label.cget("text"):
            return

        russian_table_name = title_label.cget("text").split()[1].lower()
        table_name = self.table_name_mapping.get(russian_table_name)

        if not table_name:
            return

        tree = None
        for widget in children:
            if isinstance(widget, ttk.Treeview):
                tree = widget
                break

        if tree and table_name:
            self.load_data(tree, table_name)

    # Обновление статуса заказа
    def update_order_status(self, tree):
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Внимание", "Выберите заказ для обновления статуса!")
            return

        order_id = tree.item(selected_item)['values'][0]
        current_status = tree.item(selected_item)['values'][-1]  # Последний столбец - статус

        # Окно для выбора нового статуса
        status_window = tk.Toplevel(self.root)
        status_window.title("Обновление статуса заказа")
        status_window.geometry("300x200")
        status_window.configure(bg="#f9f9f9")

        label = tk.Label(status_window, text="Текущий статус:", font=("Arial", 12), bg="#f9f9f9")
        label.pack(pady=5)

        current_status_label = tk.Label(status_window, text=current_status, font=("Arial", 12, "bold"),
                                        bg="#f9f9f9")
        current_status_label.pack(pady=5)

        new_status_label = tk.Label(status_window, text="Новый статус:", font=("Arial", 12), bg="#f9f9f9")
        new_status_label.pack(pady=5)

        status_var = tk.StringVar(value=current_status)
        status_options = ["В обработке", "Выполнен", "Отменен"]
        status_combo = ttk.Combobox(status_window, textvariable=status_var, values=status_options,
                                    font=("Arial", 12))
        status_combo.pack(pady=5)

        def save_status():
            new_status = status_var.get()
            if not new_status:
                messagebox.showerror("Ошибка", "Выберите новый статус!")
                return

            conn = sqlite3.connect('auto_parts.db')
            c = conn.cursor()
            c.execute("UPDATE orders SET status=? WHERE id=?", (new_status, order_id))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Статус заказа успешно обновлен!")
            status_window.destroy()
            self.refresh_current_view()

        save_button = tk.Button(status_window, text="Сохранить", font=("Arial", 12), bg="#4CAF50", fg="#ffffff",
                                command=save_status)
        save_button.pack(pady=10)

    # Редактирование заказа
    def edit_order(self, tree):
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Внимание", "Выберите заказ для редактирования!")
            return

        order_id = tree.item(selected_item)['values'][0]
        current_quantity = tree.item(selected_item)['values'][3]  # Количество

        edit_window = tk.Toplevel(self.root)
        edit_window.title("Редактирование заказа")
        edit_window.geometry("300x200")
        edit_window.configure(bg="#f9f9f9")

        label = tk.Label(edit_window, text="Текущее количество:", font=("Arial", 12), bg="#f9f9f9")
        label.pack(pady=5)

        quantity_label = tk.Label(edit_window, text=current_quantity, font=("Arial", 12, "bold"), bg="#f9f9f9")
        quantity_label.pack(pady=5)

        new_quantity_label = tk.Label(edit_window, text="Новое количество:", font=("Arial", 12), bg="#f9f9f9")
        new_quantity_label.pack(pady=5)

        new_quantity_entry = tk.Entry(edit_window, font=("Arial", 12))
        new_quantity_entry.pack(pady=5)

        def save_changes():
            new_quantity = new_quantity_entry.get()
            if not new_quantity:
                messagebox.showerror("Ошибка", "Введите новое количество!")
                return

            try:
                int(new_quantity)
            except ValueError:
                messagebox.showerror("Ошибка", "Количество должно быть целым числом!")
                return

            conn = sqlite3.connect('auto_parts.db')
            c = conn.cursor()
            c.execute("UPDATE orders SET quantity=? WHERE id=?", (new_quantity, order_id))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Заказ успешно обновлен!")
            edit_window.destroy()
            self.refresh_current_view()

        save_button = tk.Button(edit_window, text="Сохранить", font=("Arial", 12), bg="#4CAF50", fg="#ffffff",
                                command=save_changes)
        save_button.pack(pady=10)

    # Формирование заказа
    def create_order(self):
        self.clear_main_frame()
        frame = tk.Frame(self.main_frame, bg="#ffffff", padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        label = tk.Label(frame, text="Формирование заказа", font=("Arial", 18, "bold"), bg="#ffffff")
        label.pack(pady=10)

        supplier_label = tk.Label(frame, text="Поставщик:", font=("Arial", 12), bg="#ffffff")
        supplier_label.pack()
        supplier_combo = ttk.Combobox(frame, width=30, font=("Arial", 12))
        supplier_combo.pack()

        part_label = tk.Label(frame, text="Запчасть:", font=("Arial", 12), bg="#ffffff")
        part_label.pack()
        part_combo = ttk.Combobox(frame, width=30, font=("Arial", 12))
        part_combo.pack()

        quantity_label = tk.Label(frame, text="Количество:", font=("Arial", 12), bg="#ffffff")
        quantity_label.pack()
        quantity_entry = tk.Entry(frame, font=("Arial", 12))
        quantity_entry.pack()

        def validate_quantity():
            try:
                int(quantity_entry.get())
                return True
            except ValueError:
                messagebox.showerror("Ошибка", "Количество должно быть целым числом!")
                return False

        def save_order():
            supplier_id = supplier_combo.get().split()[0] if supplier_combo.get() else None
            part_id = part_combo.get().split()[0] if part_combo.get() else None
            quantity = quantity_entry.get()

            if not supplier_id or not part_id or not quantity:
                messagebox.showerror("Ошибка", "Заполните все поля!")
                return

            if not validate_quantity():
                return

            conn = sqlite3.connect('auto_parts.db')
            c = conn.cursor()

            # Проверяем доступное количество запчасти
            c.execute("SELECT quantity FROM parts WHERE id=?", (part_id,))
            available_quantity = c.fetchone()[0]

            if int(quantity) > available_quantity:
                messagebox.showerror("Ошибка", "Недостаточно товара на складе!")
                conn.close()
                return

            # Вычитаем количество из таблицы parts
            new_quantity = available_quantity - int(quantity)
            c.execute("UPDATE parts SET quantity=? WHERE id=?", (new_quantity, part_id))

            # Добавляем заказ
            c.execute(
                "INSERT INTO orders (supplier_id, part_id, quantity, order_date, status) VALUES (?, ?, ?, datetime('now'), ?)",
                (supplier_id, part_id, quantity, "В обработке")
            )
            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Заказ успешно создан!")
            self.clear_main_frame()
            self.show_ref("orders")

        save_button = tk.Button(frame, text="Создать заказ", font=("Arial", 12), bg="#4CAF50", fg="#ffffff",
                                command=save_order)
        save_button.pack(pady=10)

        # Загрузка данных в Combobox
        conn = sqlite3.connect('auto_parts.db')
        c = conn.cursor()
        c.execute("SELECT id, name FROM suppliers")
        suppliers = [f"{id} {name}" for id, name in c.fetchall()]
        supplier_combo['values'] = suppliers

        c.execute("SELECT id, name FROM parts")
        parts = [f"{id} {name}" for id, name in c.fetchall()]
        part_combo['values'] = parts
        conn.close()

    # Показать отчет о продажах
    def show_sales_report(self):
        self.clear_main_frame()
        frame = tk.Frame(self.main_frame, bg="#ffffff", padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        label = tk.Label(frame, text="Отчет о продажах", font=("Arial", 18, "bold"), bg="#ffffff")
        label.pack(pady=10)

        tree = ttk.Treeview(frame, columns=("ID", "Client", "Part", "Quantity", "Date"), show="headings", height=15,
                            style="Custom.Treeview")
        tree.heading("ID", text="ID")
        tree.heading("Client", text="Клиент")
        tree.heading("Part", text="Запчасть")
        tree.heading("Quantity", text="Количество")
        tree.heading("Date", text="Дата")
        tree.pack(fill=tk.BOTH, expand=True)

        self.load_sales(tree)

    # Загрузить данные продаж в Treeview
    def load_sales(self, tree):
        conn = sqlite3.connect('auto_parts.db')
        c = conn.cursor()
        c.execute("""
                SELECT sales.id, clients.name, parts.name, sales.quantity, sales.sale_date
                FROM sales
                INNER JOIN clients ON sales.client_id = clients.id
                INNER JOIN parts ON sales.part_id = parts.id
            """)
        rows = c.fetchall()
        for row in rows:
            tree.insert("", tk.END, values=row)
        conn.close()

    # Показать отчет о заказах
    def show_orders_report(self):
        self.clear_main_frame()
        frame = tk.Frame(self.main_frame, bg="#ffffff", padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        label = tk.Label(frame, text="Отчет о заказах", font=("Arial", 18, "bold"), bg="#ffffff")
        label.pack(pady=10)

        tree = ttk.Treeview(frame, columns=("ID", "Supplier", "Part", "Quantity", "Date", "Status"),
                            show="headings", height=15, style="Custom.Treeview")
        tree.heading("ID", text="ID")
        tree.heading("Supplier", text="Поставщик")
        tree.heading("Part", text="Запчасть")
        tree.heading("Quantity", text="Количество")
        tree.heading("Date", text="Дата")
        tree.heading("Status", text="Статус")
        tree.pack(fill=tk.BOTH, expand=True)

        self.load_orders(tree)

    # Загрузить данные заказов в Treeview
    def load_orders(self, tree):
        conn = sqlite3.connect('auto_parts.db')
        c = conn.cursor()
        c.execute("""
                SELECT orders.id, suppliers.name, parts.name, orders.quantity, orders.order_date, orders.status
                FROM orders
                INNER JOIN suppliers ON orders.supplier_id = suppliers.id
                INNER JOIN parts ON orders.part_id = parts.id
            """)
        rows = c.fetchall()
        for row in rows:
            tree.insert("", tk.END, values=row)
        conn.close()

    # Показать руководство
    def show_manual(self):
        self.clear_main_frame()
        frame = tk.Frame(self.main_frame, bg="#ffffff", padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        label = tk.Label(frame, text="Руководство по использованию", font=("Arial", 18, "bold"), bg="#ffffff")
        label.pack(pady=10)

        manual_text = """
            1. Используйте меню 'Справочники' для просмотра, добавления, редактирования и удаления записей.
            2. В разделе 'Отчеты' вы можете увидеть информацию о продажах и заказах.
            3. Для создания заказа используйте пункт меню 'Создать заказ'.
            """
        manual_label = tk.Label(frame, text=manual_text, justify="left", font=("Arial", 12), bg="#ffffff")
        manual_label.pack(pady=10)

    # Очистить основной фрейм
    def clear_main_frame(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()

# Создание базы данных и запуск приложения
if __name__ == "__main__":
    create_database()
    root = tk.Tk()
    app = AutoPartsApp(root)
    root.mainloop()
